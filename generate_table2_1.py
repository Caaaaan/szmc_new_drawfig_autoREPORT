import os
import re
import pandas as pd

OUTPUT_NAME = "table2_1.xlsx"


def _parse_date_from_filename(fname: str):
    """从文件名中提取日期字符串 YYYY年MM月DD日"""
    m = re.search(r"(\d{4}年\d{2}月\d{2}日)", fname)
    return m.group(1) if m else ""


def _parse_line_name(fname: str) -> str:
    """从文件名中提取线路名称，如 '深圳地铁11号线' 或 '深圳1号线'"""
    # 兼容两种格式: "深圳地铁N号线" 和 "深圳N号线"
    m = re.search(r"全线路原始数据_(深圳(?:地铁)?\d+号线)_", fname)
    return m.group(1) if m else ""


def _parse_line_number(line_name: str):
    """从线路名称中提取线路数字编号，如 '深圳地铁11号线' -> 11"""
    m = re.search(r"(\d+)号线", line_name)
    return int(m.group(1)) if m else None


def _find_latest_files(data_dir: str):
    """找到 original_data 下最新日期的上下行两个文件路径。
    优先扫描 上行/下行 子目录；子目录不存在时回退到 flat 目录（兼容旧任务）。

    Returns: (up_path, down_path, date_str)
    """
    up_dir = os.path.join(data_dir, "上行")
    down_dir = os.path.join(data_dir, "下行")

    if os.path.isdir(up_dir) and os.path.isdir(down_dir):
        scan_plan = [(up_dir, "上行"), (down_dir, "下行")]
    else:
        scan_plan = [(data_dir, None)]

    dated_up = {}
    dated_down = {}
    for scan_dir, known_dir in scan_plan:
        if not os.path.isdir(scan_dir):
            continue
        for f in sorted(os.listdir(scan_dir)):
            if not f.endswith(".xlsx"):
                continue
            d = _parse_date_from_filename(f)
            if not d:
                continue
            if known_dir == "上行" or (known_dir is None and "上行" in f):
                dated_up.setdefault(d, []).append(os.path.join(scan_dir, f))
            elif known_dir == "下行" or (known_dir is None and "下行" in f):
                dated_down.setdefault(d, []).append(os.path.join(scan_dir, f))

    all_dates = set(dated_up.keys()) | set(dated_down.keys())
    if not all_dates:
        raise ValueError("无法从文件名中解析日期")

    latest_date = max(all_dates)
    up_path = dated_up.get(latest_date, [None])[0]
    down_path = dated_down.get(latest_date, [None])[0]

    print(f"最新数据日期: {latest_date}")
    print(f"  上行: {os.path.basename(up_path) if up_path else '无'}")
    print(f"  下行: {os.path.basename(down_path) if down_path else '无'}")

    return up_path, down_path, latest_date


def _grade_height_diff(diff: float) -> str:
    """导高驰度分级: 按(站区+杆号)组内高差"""
    if diff > 25 :
        return "Ⅲ"
    if 15 < diff <= 25:
        return "Ⅱ"
    if 10 < diff <= 15:
        return "Ⅰ"
    return ""


def _grade_wear_width(val: float) -> str:
    """磨耗宽度分级（标准标准）"""
    if val > 13:
        return "Ⅲ"
    if 10 < val <= 13:
        return "Ⅱ"
    if 8 < val <= 10:
        return "Ⅰ"
    return ""


def _grade_wear_width_line1_5(val: float) -> str:
    """磨耗宽度分级（1号线/5号线专用）"""
    if val > 11:
        return "Ⅲ"
    if 8 < val <= 11:
        return "Ⅱ"
    if 7 < val <= 8:
        return "Ⅰ"
    return ""


def _grade_pressure_low(val: float) -> str:
    """接触压力 <100N 分级 接触压力最低有效值，<=10N 认为异常"""
    if 10<= val < 50:
        return "Ⅲ"
    if 50 <= val < 80:
        return "Ⅱ"
    if 80 <= val < 100:
        return "Ⅰ"
    return ""


def _grade_pressure_high(val: float) -> str:
    """接触压力 >130N 分级"""
    if val > 220:
        return "Ⅲ"
    if 180 < val <= 220:
        return "Ⅱ"
    if 130 < val <= 180:
        return "Ⅰ"
    return ""


def _grade_hardpoint(val: float) -> str:
    """硬点分级 (硬点2 m/s^2)（标准标准）"""
    if val > 200:
        return "Ⅲ"
    if 130 < val <= 200:
        return "Ⅱ"
    if 70 < val <= 130:
        return "Ⅰ"
    return ""


def _grade_hardpoint_line11(val: float) -> str:
    """硬点分级 (硬点2 m/s^2)（11号线专用）"""
    if val > 250:
        return "Ⅲ"
    if 150 < val <= 250:
        return "Ⅱ"
    if 100 < val <= 150:
        return "Ⅰ"
    return ""


def compute_defect_counts(df_up: pd.DataFrame, df_down: pd.DataFrame, line_name: str = "") -> dict:
    """合并上下行数据，计算各缺陷等级个数。

    Parameters
    ----------
    line_name : str
        线路名称（如 '深圳地铁11号线'）。仅当线路为11号线或5号线时，
        导高驰度和磨耗宽度的计算会过滤 导高 < 4400mm 的数据点。
    """
    line_num = _parse_line_number(line_name)
    _filter_height_4400 = line_num in (5, 11)  # 5号线/11号线需过滤导高<4400mm

    # ---------- 导高驰度 ----------
    height_diffs = []
    for df_dir in [df_up, df_down]:
        if "导高(mm)" not in df_dir.columns:
            continue
        sub = df_dir
        # 11号线/5号线仅统计导高 < 4400mm 的数据点
        if _filter_height_4400:
            sub = sub[sub["导高(mm)"] < 4400]
        if len(sub) == 0:
            continue
        # groupby 与 draw_process.py graph5 对齐：用字符串拼接键
        grp_key = sub["站区"].astype(str) + "_" + sub["杆号"].astype(str)
        height_range = sub.groupby(grp_key)["导高(mm)"].agg(["max", "min"])
        diffs = height_range["max"] - height_range["min"]
        height_diffs.append(diffs.dropna())
    if height_diffs:
        all_hd = pd.concat(height_diffs, ignore_index=True)
        hd_grades = all_hd.apply(_grade_height_diff)
        hd_counts = hd_grades.value_counts()
    else:
        hd_counts = pd.Series(dtype=int)

    # ---------- 磨耗宽度 ----------
    wear_widths = []
    for df_dir in [df_up, df_down]:
        if "磨耗宽度(mm)" in df_dir.columns:
            sub = df_dir.dropna(subset=["磨耗宽度(mm)"])
            # 11号线/5号线仅统计导高 < 4400mm 的数据点
            if _filter_height_4400 and "导高(mm)" in df_dir.columns:
                sub = sub[sub["导高(mm)"] < 4400]
            if len(sub) > 0:
                wear_widths.append(sub["磨耗宽度(mm)"])
    if wear_widths:
        all_ww = pd.concat(wear_widths, ignore_index=True)
        ww_grades = all_ww.apply(_grade_wear_width_line1_5 if (line_num in (1, 5)) else _grade_wear_width)
        ww_counts = ww_grades.value_counts()
    else:
        ww_counts = pd.Series(dtype=int)

    # ---------- 接触压力 ----------
    pressures = []
    for df_dir in [df_up, df_down]:
        pres_col = None
        for candidate in ['压力和(N)', '压力和']:
            if candidate in df_dir.columns:
                pres_col = candidate
                break
        if pres_col:
            pressures.append(df_dir[pres_col].dropna())
    if pressures:
        all_pr = pd.concat(pressures, ignore_index=True)
        all_pr = all_pr[all_pr > 10]  # <=10N视为接触异常，不予统计（与draw_process对齐）
        pr_low_grades = all_pr.apply(_grade_pressure_low)
        pr_low_counts = pr_low_grades.value_counts()
        pr_high_grades = all_pr.apply(_grade_pressure_high)
        pr_high_counts = pr_high_grades.value_counts()
    else:
        pr_low_counts = pd.Series(dtype=int)
        pr_high_counts = pd.Series(dtype=int)

    # ---------- 硬点 ----------
    hardpoint_vals = []
    hardpoint_cols = ['硬点1(m/s²)', '硬点2(m/s²)']
    for df_dir in [df_up, df_down]:
        available = [c for c in hardpoint_cols if c in df_dir.columns]
        if not available:
            continue
        # 与draw_process对齐：硬点1或硬点2 > 7g 即算超限（OR逻辑）
        mask = pd.Series(False, index=df_dir.index)
        for col in available:
            series = pd.to_numeric(df_dir[col], errors='coerce')
            mask = mask | (series > 70)
        # 取该行所有硬点列的最大值用于分级
        row_max = df_dir[available].apply(pd.to_numeric, errors='coerce').max(axis=1)
        hardpoint_vals.append(row_max.loc[mask].dropna())
    if hardpoint_vals:
        all_hp = pd.concat(hardpoint_vals, ignore_index=True)
        hp_grades = all_hp.apply(_grade_hardpoint_line11 if (line_num == 11) else _grade_hardpoint)
        hp_counts = hp_grades.value_counts()
    else:
        hp_counts = pd.Series(dtype=int)

    return {
        "导高驰度": hd_counts,
        "磨耗宽度": ww_counts,
        "接触压力_低": pr_low_counts,
        "接触压力_高": pr_high_counts,
        "硬点": hp_counts,
    }


def _safe(series: pd.Series, grade: str) -> int:
    return int(series.get(grade, 0))


def write_result_excel(counts: dict, outpath: str):
    """直接写入带合并单元格格式的 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Side, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "几何参数缺陷统计"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ---- 表头 (7 列) ----
    # 类别 | 问题类型(colspan=2) | 单位 | Ⅰ级 | Ⅱ级 | Ⅲ级
    headers = [
        (1, "类别"),
        (2, "问题类型"),
        (4, "单位"),
        (5, "Ⅰ级"),
        (6, "Ⅱ级"),
        (7, "Ⅲ级"),
    ]
    for c, h in headers:
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = center_align
        cell.border = thin_border
    # 表头 B1:C1 合并
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=3)

    # ---- 数据行 (7 列: A=类别, B=问题类型, C=子类型, D=单位, E=Ⅰ, F=Ⅱ, G=Ⅲ) ----
    data_rows = [
        ["几何参数缺陷", "导高驰度(mm)", None, "处",
         _safe(counts["导高驰度"], "Ⅰ"),
         _safe(counts["导高驰度"], "Ⅱ"),
         _safe(counts["导高驰度"], "Ⅲ")],
        ["几何参数缺陷", "磨耗宽度（mm）", None, "处",
         _safe(counts["磨耗宽度"], "Ⅰ"),
         _safe(counts["磨耗宽度"], "Ⅱ"),
         _safe(counts["磨耗宽度"], "Ⅲ")],
        ["几何参数缺陷", "接触压力(＜100N)", None, "处",
         _safe(counts["接触压力_低"], "Ⅰ"),
         _safe(counts["接触压力_低"], "Ⅱ"),
         _safe(counts["接触压力_低"], "Ⅲ")],
        ["几何参数缺陷", "接触压力(＞130N)", None, "处",
         _safe(counts["接触压力_高"], "Ⅰ"),
         _safe(counts["接触压力_高"], "Ⅱ"),
         _safe(counts["接触压力_高"], "Ⅲ")],
        ["几何参数缺陷", "硬点（g）", None, "处",
         _safe(counts["硬点"], "Ⅰ"),
         _safe(counts["硬点"], "Ⅱ"),
         _safe(counts["硬点"], "Ⅲ")],
    ]

    for r, row_data in enumerate(data_rows, 2):
        for c, val in enumerate(row_data, 1):
            if val is None:
                continue
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = center_align
            cell.border = thin_border

    # ---- 合并单元格 ----
    # 类别: A2:A6
    ws.merge_cells(start_row=2, start_column=1, end_row=6, end_column=1)
    # 导高驰度: B2:C2
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=3)
    # 磨耗宽度: B3:C3
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=3)
    # 接触压力(＜100N): B4:C4
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=3)
    # 接触压力(＞130N): B5:C5
    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=3)
    # 硬点: B6:C6
    ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=3)

    # ---- 列宽 ----
    col_widths = {"A": 16, "B": 20, "C": 12, "D": 8, "E": 10, "F": 10, "G": 10}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"

    os.makedirs(os.path.dirname(outpath) or ".", exist_ok=True)
    wb.save(outpath)
    print(f"已保存: {outpath}")


def run(data_dir: str = "original_data", output_dir: str = "output_excel"):
    """供 Web 服务调用：从 data_dir 加载最新日期的上下行文件，
    计算几何参数缺陷统计并写入 table2_1.xlsx。

    所有路径均已参数化，支持多用户并发。
    """
    up_path, down_path, date_str = _find_latest_files(data_dir)

    df_up = pd.DataFrame()
    df_down = pd.DataFrame()

    if up_path:
        df_up = pd.read_excel(up_path)
        before = len(df_up)
        df_up = df_up.dropna(subset=['站区'])
        if len(df_up) < before:
            print(f"  已过滤 {before - len(df_up)} 行无效数据（站区为空）")
        print(f"  加载 {os.path.basename(up_path)}: {len(df_up)} 行 (上行)")

    if down_path:
        df_down = pd.read_excel(down_path)
        before = len(df_down)
        df_down = df_down.dropna(subset=['站区'])
        if len(df_down) < before:
            print(f"  已过滤 {before - len(df_down)} 行无效数据（站区为空）")
        print(f"  加载 {os.path.basename(down_path)}: {len(df_down)} 行 (下行)")

    line_name = ""
    for p in [up_path, down_path]:
        if p:
            line_name = _parse_line_name(os.path.basename(p))
            break
    if line_name:
        print(f"  识别线路: {line_name}")

    counts = compute_defect_counts(df_up, df_down, line_name)

    print("\n缺陷统计结果:")
    labels = {
        "导高驰度": "导高驰度",
        "磨耗宽度": "磨耗宽度",
        "接触压力_低": "接触压力<100N",
        "接触压力_高": "接触压力>130N",
        "硬点": "硬点",
    }
    for k, label in labels.items():
        v = counts[k]
        print(f"  {label}: Ⅰ={_safe(v, 'Ⅰ')}, Ⅱ={_safe(v, 'Ⅱ')}, Ⅲ={_safe(v, 'Ⅲ')}")

    outpath = os.path.join(output_dir, OUTPUT_NAME)
    write_result_excel(counts, outpath)
    print(f"\n表格生成完毕: {outpath}")


def main():
    run("original_data", "output_excel")


if __name__ == "__main__":
    main()

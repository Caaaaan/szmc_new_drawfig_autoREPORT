"""
根据 default_report 文件夹下的内容，调用 DeepSeek API 进行数据整合，
生成 table2_2.xlsx（各等级情况总览）和 table2_3.xlsx（项点情况总览）。
"""

import os
import re
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openai import OpenAI

# ---------------------------------------------------------------------------
# 1. 数据读取
# ---------------------------------------------------------------------------

def read_all_excel_files(data_dir: str) -> dict:
    """读取 data_dir 下所有 xlsx 文件，返回 {文件名: DataFrame}。"""
    if not os.path.isdir(data_dir):
        raise FileNotFoundError(f"数据目录不存在: {data_dir}")

    files = sorted([f for f in os.listdir(data_dir) if f.endswith(".xlsx")])
    if not files:
        raise FileNotFoundError(f"在 {data_dir} 中未找到 Excel 文件")

    all_data = {}
    for fname in files:
        fpath = os.path.join(data_dir, fname)
        try:
            df = pd.read_excel(fpath)
            all_data[fname] = df
            print(f"  已加载: {fname} ({len(df)} 行, {len(df.columns)} 列)")
        except Exception as e:
            print(f"  警告: 读取 {fname} 失败: {e}")

    if not all_data:
        raise ValueError("未能成功读取任何 Excel 文件")
    return all_data


def build_data_summary(all_data: dict, max_rows_per_file: int = 100) -> str:
    """将多个 DataFrame 拼接为文本摘要，供 API 调用时使用。"""
    parts = []
    for fname, df in all_data.items():
        parts.append(f"=== 文件: {fname} ===")
        parts.append(f"列名: {list(df.columns)}")
        parts.append(f"行数: {len(df)}")
        # 截断过长的表格
        tail = df.head(max_rows_per_file).to_string(index=False)
        parts.append(tail)
        if len(df) > max_rows_per_file:
            parts.append(f"... (共 {len(df)} 行，仅展示前 {max_rows_per_file} 行)")
        parts.append("")
    return "\n".join(parts)


# ---------------------------------------------------------------------------
# 2. DeepSeek API 调用
# ---------------------------------------------------------------------------

def call_deepseek_api(data_summary: str) -> dict:
    """调用 DeepSeek API，让模型从缺陷数据中提取结构化统计。"""
    api_key = os.environ.get("DEEPSEEK_API_KEY")
    if not api_key:
        raise RuntimeError("未设置环境变量 DEEPSEEK_API_KEY")

    client = OpenAI(
        api_key=api_key,
        base_url="https://api.deepseek.com",
    )

    system_prompt = (
        "你是一个铁路接触网检测数据分析专家。"
        "你的任务是根据提供的缺陷/检测数据，输出结构化的统计 JSON。"
        "请严格按照要求的 JSON 格式输出，不要输出任何额外的解释文字。"
        "注意：缺陷位置(location)应填写设备部件类别（如'绝缘子''绝缘锚段''刚性支持定位装置'"
        "'刚性绝缘类''接触悬挂类'等），而不是具体站区名称。"
    )

    user_prompt = f"""请根据以下缺陷检测数据，生成两个统计表格的数据。

数据内容：
{data_summary}

请分析以上数据，提取所有缺陷记录，并按照以下 JSON 格式返回：

{{
  "table2_2": [
    {{"grade": "一级", "count": <数量>}},
    {{"grade": "二级", "count": <数量>}},
    {{"grade": "三级", "count": <数量>}}
  ],
  "table2_3": [
    {{
      "location": "<缺陷位置>",
      "description": "<缺陷描述>",
      "grade": "<缺陷级别>",
      "count": <汇总数量>
    }}
  ]
}}

要求：
1. table2_2 中的三个等级都必须存在，即使数量为 0。
2. table2_3 按缺陷位置+缺陷描述+缺陷级别分组汇总，同一类缺陷合并计数。
3. 缺陷级别使用中文："一级"、"二级"、"三级"（一级最严重）。
4. **缺陷位置(location)必须填写设备部件类别**，如"绝缘子""绝缘锚段""刚性支持定位装置"
   "刚性绝缘类""接触悬挂类"等，**严禁填写具体站区名或区间名**。
5. 如果数据中包含明确的缺陷位置、缺陷描述、缺陷级别列，直接使用；如果数据是自然语言描述，
   请智能提取分类。
6. 只输出 JSON，不要包含任何 markdown 代码块标记或其他文字。
"""

    print("\n正在调用 DeepSeek API 进行数据整合分析...")
    response = client.chat.completions.create(
        model="deepseek-v4-pro",
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        stream=False,
        reasoning_effort="high",
        extra_body={"thinking": {"type": "enabled"}},
    )

    content = response.choices[0].message.content.strip()
    print(f"API 原始响应 (前 500 字符):\n{content[:500]}")

    # 尝试清理可能的 markdown 代码块包裹
    content = re.sub(r"^```(?:json)?\s*", "", content)
    content = re.sub(r"\s*```$", "", content)

    try:
        result = json.loads(content)
    except json.JSONDecodeError as e:
        print(f"JSON 解析失败: {e}")
        print(f"完整响应:\n{content}")
        raise

    # 基本校验
    if "table2_2" not in result or "table2_3" not in result:
        raise ValueError("API 返回的 JSON 缺少 table2_2 或 table2_3 字段")

    return result


# ---------------------------------------------------------------------------
# 3. Excel 写入
# ---------------------------------------------------------------------------

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BOLD_FONT = Font(bold=True)


def _style_cell(ws, row, col, value, bold=False):
    """设置单元格值和样式。"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    if bold:
        cell.font = BOLD_FONT
    return cell


def write_table2_2(data: list, outpath: str):
    """写入 表 2-2 各等级情况总览。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "各等级情况总览"

    # 表头
    headers = ["序号", "缺陷等级", "确认缺陷数量"]
    for c, h in enumerate(headers, 1):
        _style_cell(ws, 1, c, h, bold=True)

    # 按 一级→二级→三级 固定顺序
    grade_order = {"一级": 1, "二级": 2, "三级": 3}
    rows = sorted(data, key=lambda x: grade_order.get(x.get("grade", ""), 99))

    total = 0
    for i, item in enumerate(rows):
        r = i + 2
        grade = item.get("grade", "")
        count = int(item.get("count", 0))
        total += count
        _style_cell(ws, r, 1, i + 1)
        _style_cell(ws, r, 2, grade)
        _style_cell(ws, r, 3, count)

    # 合计行
    total_row = len(rows) + 2
    _style_cell(ws, total_row, 1, "合计", bold=True)
    _style_cell(ws, total_row, 2, "")
    _style_cell(ws, total_row, 3, total, bold=True)

    # 列宽
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16

    ws.freeze_panes = "A2"
    os.makedirs(os.path.dirname(outpath), exist_ok=True)
    wb.save(outpath)
    print(f"已保存: {outpath} (共 {total} 项缺陷)")


def write_table2_3(data: list, outpath: str):
    """写入 表 2-3 项点情况总览。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "项点情况总览"

    # 表头
    headers = ["序号", "缺陷位置", "缺陷描述", "缺陷级别", "汇总"]
    for c, h in enumerate(headers, 1):
        _style_cell(ws, 1, c, h, bold=True)

    total = 0
    for i, item in enumerate(data):
        r = i + 2
        location = item.get("location", "")
        description = item.get("description", "")
        grade = item.get("grade", "")
        count = int(item.get("count", 0))
        total += count
        _style_cell(ws, r, 1, i + 1)
        _style_cell(ws, r, 2, location)
        _style_cell(ws, r, 3, description)
        _style_cell(ws, r, 4, grade)
        _style_cell(ws, r, 5, count)

    # 合计行
    total_row = len(data) + 2
    _style_cell(ws, total_row, 1, "合计", bold=True)
    _style_cell(ws, total_row, 2, "")
    _style_cell(ws, total_row, 3, "")
    _style_cell(ws, total_row, 4, "")
    _style_cell(ws, total_row, 5, total, bold=True)

    # 列宽
    col_widths = {"A": 8, "B": 22, "C": 40, "D": 12, "E": 10}
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width

    ws.freeze_panes = "A2"
    os.makedirs(os.path.dirname(outpath), exist_ok=True)
    wb.save(outpath)
    print(f"已保存: {outpath} (共 {len(data)} 类缺陷, {total} 项)")


# ---------------------------------------------------------------------------
# 4. 主流程
# ---------------------------------------------------------------------------

def run(data_dir: str = "defect_report", output_dir: str = "output_excel"):
    """供 Web 服务调用：从 data_dir 读取缺陷报告，调用 DeepSeek API 整合，

    输出 table2_2.xlsx 和 table2_3.xlsx 到 output_dir。

    所有路径均已参数化，支持多用户并发。
    """
    print("=" * 60)
    print("生成 table2_2.xlsx 和 table2_3.xlsx")
    print("=" * 60)

    # 1. 读取数据
    print(f"\n[1/3] 读取 {data_dir}/ 目录下的数据文件...")
    all_data = read_all_excel_files(data_dir)
    print(f"  共读取 {len(all_data)} 个文件")

    # 2. 调用 API
    print("\n[2/3] 调用 DeepSeek API 进行数据整合...")
    data_summary = build_data_summary(all_data)
    result = call_deepseek_api(data_summary)

    # 3. 输出 Excel
    print("\n[3/3] 生成 Excel 表格...")
    outpath_2_2 = os.path.join(output_dir, "table2_2.xlsx")
    outpath_2_3 = os.path.join(output_dir, "table2_3.xlsx")
    write_table2_2(result["table2_2"], outpath_2_2)
    write_table2_3(result["table2_3"], outpath_2_3)

    print("\n" + "=" * 60)
    print("表格生成完毕！")
    print(f"  {outpath_2_2}")
    print(f"  {outpath_2_3}")
    print("=" * 60)


def main():
    run("defect_report", "output_excel")


if __name__ == "__main__":
    main()

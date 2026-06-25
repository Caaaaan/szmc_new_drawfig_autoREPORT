"""
根据 original_data 中最新的原始数据，按上下行分别整理生成两个汇总数据表：
  数据表1 — 按拉出值区间统计：拉出值区间、拉出值次数、压力＜110N压力＞130N（次数）、硬点大于7g/10g（次数）、磨耗大于8mm（次数）
  数据表2 — 按站点统计：站点、压力＜110N压力＞130N（次数）、硬点大于7g/10g（次数）、磨耗大于8mm（次数）
"""

import os
import re
import json
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font

from draw_process import needs_height_filter

# 阈值常量（与 draw_process.py 保持一致）
HARDPOINT_THRESHOLD = 7 * 10   # 7g, g=10m/s² → 70 m/s²
PRESSURE_LOW_THRESHOLD = 110   # N
PRESSURE_HIGH_THRESHOLD = 130  # N
WEAR_WIDTH_THRESHOLD = 8       # mm
HEIGHT_MAX_THRESHOLD = 4400    # mm（特定线路导高上限）

# 拉出值区间定义（与 draw_process.py 保持一致）
PULLOUT_RANGE_START = -250
PULLOUT_RANGE_END = 250
PULLOUT_BIN_WIDTH = 10

PULLOUT_BIN_EDGES = list(range(PULLOUT_RANGE_START, PULLOUT_RANGE_END + 1, PULLOUT_BIN_WIDTH))
PULLOUT_BIN_LABELS = [
    f"{PULLOUT_BIN_EDGES[i]}-{PULLOUT_BIN_EDGES[i + 1]}"
    for i in range(len(PULLOUT_BIN_EDGES) - 1)
]
PULLOUT_BIN_LABELS = ["<-250"] + PULLOUT_BIN_LABELS + [">250"]
PULLOUT_BINS = [-float("inf")] + PULLOUT_BIN_EDGES + [float("inf")]

# ---------------------------------------------------------------------------
# 辅助函数
# ---------------------------------------------------------------------------

def _parse_date_from_filename(fname: str) -> str:
    m = re.search(r"(\d{4}年\d{2}月\d{2}日)", fname)
    return m.group(1) if m else ""


def _find_latest_files(data_dir: str) -> tuple:
    """找到 original_data 下最新日期的上下行两个文件路径。
    优先扫描 上行/下行 子目录；子目录不存在时回退到 flat 目录（兼容旧任务）。

    Returns: (up_path, down_path, date_str)
    """
    up_dir = os.path.join(data_dir, "上行")
    down_dir = os.path.join(data_dir, "下行")

    if os.path.isdir(up_dir) and os.path.isdir(down_dir):
        scan_plan = [(up_dir, "上行"), (down_dir, "下行")]
    else:
        scan_plan = [(data_dir, None)]

    dated_up = {}
    dated_down = {}
    for scan_dir, known_dir in scan_plan:
        if not os.path.isdir(scan_dir):
            continue
        for f in sorted(os.listdir(scan_dir)):
            if not f.endswith(".xlsx"):
                continue
            d = _parse_date_from_filename(f)
            if not d:
                continue
            if known_dir == "上行" or (known_dir is None and "上行" in f):
                dated_up.setdefault(d, []).append(os.path.join(scan_dir, f))
            elif known_dir == "下行" or (known_dir is None and "下行" in f):
                dated_down.setdefault(d, []).append(os.path.join(scan_dir, f))

    all_dates = set(dated_up.keys()) | set(dated_down.keys())
    if not all_dates:
        raise ValueError("无法从文件名中解析日期")

    latest_date = max(all_dates)
    up_path = dated_up.get(latest_date, [None])[0]
    down_path = dated_down.get(latest_date, [None])[0]

    print(f"最新数据日期: {latest_date}")
    print(f"  上行: {os.path.basename(up_path) if up_path else '无'}")
    print(f"  下行: {os.path.basename(down_path) if down_path else '无'}")

    return up_path, down_path, latest_date


def _get_hardpoint_mask(df: pd.DataFrame) -> pd.Series:
    """硬点 > 7g (任一硬点列 > 70 m/s²)"""
    candidates = ["硬点1(m/s²)", "硬点2(m/s²)"]
    available = [c for c in candidates if c in df.columns]
    mask = pd.Series(False, index=df.index)
    for col in available:
        mask = mask | (pd.to_numeric(df[col], errors="coerce") > HARDPOINT_THRESHOLD)
    return mask


def _get_pressure_out_of_range_mask(df: pd.DataFrame) -> pd.Series:
    """压力 < 110N 或 压力 > 130N"""
    for candidate in ["压力和(N)", "压力和"]:
        if candidate in df.columns:
            vals = pd.to_numeric(df[candidate], errors="coerce")
            return (vals < PRESSURE_LOW_THRESHOLD) | (vals > PRESSURE_HIGH_THRESHOLD)
    return pd.Series(False, index=df.index)


def _get_wear_width_mask(df: pd.DataFrame, line_number: str = "") -> pd.Series:
    """磨耗宽度 > 8mm（特定线路时排除导高 > 4400mm 的数据点）"""
    if "磨耗宽度(mm)" not in df.columns:
        return pd.Series(False, index=df.index)

    mask = pd.to_numeric(df["磨耗宽度(mm)"], errors="coerce") > WEAR_WIDTH_THRESHOLD

    # 特定线路：导高大于4400mm的数据点不统计
    if needs_height_filter(line_number) and "导高(mm)" in df.columns:
        height_mask = pd.to_numeric(df["导高(mm)"], errors="coerce") < HEIGHT_MAX_THRESHOLD
        mask = mask & height_mask

    return mask


def _apply_pullout_bin(pullout: pd.Series) -> pd.Series:
    return pd.cut(pullout, bins=PULLOUT_BINS, labels=PULLOUT_BIN_LABELS, right=False)


def _load_station_order(line_number: str = "", direction: str = "",
                        line_name: str = "",
                        station_json_path: str = "station.json") -> list:
    """从 station.json 中读取指定线路和行别的站区顺序。

    优先使用 line_number（来自 welcome.html）+ direction 精确匹配；
    回退到 line_name（数据列值）模糊匹配。
    找不到则返回空列表。
    """
    try:
        with open(station_json_path, "r", encoding="utf-8") as f:
            station_data = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        print(f"  警告: 未找到或无法解析 {station_json_path}")
        return []

    # 优先：使用用户交互提供的 line_number 精确匹配
    if line_number and direction:
        target_suffix = f"-{direction}"
        for skey in station_data:
            if line_number in skey and skey.endswith(target_suffix):
                return station_data[skey]
        print(f"  警告: 未找到线路'{line_number}'行别'{direction}'的站区顺序")

    # 回退：使用数据中的 line_name 模糊匹配（兼容 CLI 模式）
    if line_name and direction:
        key = f"{line_name}-{direction}"
        if key in station_data:
            return station_data[key]
        for skey in station_data:
            if skey.startswith(line_name + "-") or line_name.startswith(skey.split("-")[0]):
                if skey.endswith("-" + direction):
                    print(f"  模糊匹配: 数据线名'{line_name}' -> station.json键'{skey}'")
                    return station_data[skey]
        print(f"  警告: 未找到线名'{line_name}'行别'{direction}'的站区顺序")

    return []


def _build_ordered_station_list(df: pd.DataFrame, direction: str,
                                station_json_path: str = "station.json",
                                line_number: str = "") -> list:
    """将 station.json 的顺序与数据中实际出现的站区合并，返回有序站区列表。"""
    # 回退：从数据列获取 line_name（兼容 CLI 模式无 line_number 的情况）
    line_name = df["线名"].iloc[0] if "线名" in df.columns and len(df) > 0 else ""
    if not line_number and not line_name:
        return sorted(df["站区"].dropna().unique())

    ordered = _load_station_order(line_number=line_number, direction=direction,
                                  line_name=line_name,
                                  station_json_path=station_json_path)
    ordered_set = set(ordered)

    extra = set()
    for s in df["站区"].dropna().unique():
        if s not in ordered_set and isinstance(s, str):
            extra.add(s)

    return ordered + sorted(extra)


# ---------------------------------------------------------------------------
# 构建数据表（单方向）
# ---------------------------------------------------------------------------

def build_table1(df: pd.DataFrame, line_number: str = "") -> pd.DataFrame:
    """数据表1：按拉出值区间统计"""
    pullout_raw = pd.to_numeric(df["拉出值(mm)"], errors="coerce")
    valid = pullout_raw.notna()
    pullout = pullout_raw[valid]
    df_valid = df.loc[valid]

    bin_series = _apply_pullout_bin(pullout)
    total_counts = bin_series.value_counts()
    pressure_counts = bin_series[_get_pressure_out_of_range_mask(df_valid)].value_counts()
    hardpoint_counts = bin_series[_get_hardpoint_mask(df_valid)].value_counts()
    wear_counts = bin_series[_get_wear_width_mask(df_valid, line_number=line_number)].value_counts()

    return pd.DataFrame({
        "拉出值区间": PULLOUT_BIN_LABELS,
        "拉出值次数":     [int(total_counts.get(lbl, 0)) for lbl in PULLOUT_BIN_LABELS],
        "压力＜110N压力＞130N（次数）":  [int(pressure_counts.get(lbl, 0)) for lbl in PULLOUT_BIN_LABELS],
        "硬点大于7g/10g（次数）": [int(hardpoint_counts.get(lbl, 0)) for lbl in PULLOUT_BIN_LABELS],
        "磨耗大于8mm（次数）":   [int(wear_counts.get(lbl, 0)) for lbl in PULLOUT_BIN_LABELS],
    })


def build_table2(df: pd.DataFrame, station_order: list,
                 line_number: str = "") -> pd.DataFrame:
    """数据表2：按站点统计，行序按 station_order 排列。"""
    df_valid = df.dropna(subset=["站区"])

    pressure_counts = df_valid[_get_pressure_out_of_range_mask(df_valid)]["站区"].value_counts()
    hardpoint_counts = df_valid[_get_hardpoint_mask(df_valid)]["站区"].value_counts()
    wear_counts = df_valid[_get_wear_width_mask(df_valid, line_number=line_number)]["站区"].value_counts()

    return pd.DataFrame({
        "站点": station_order,
        "压力＜110N压力＞130N（次数）":  [int(pressure_counts.get(s, 0)) for s in station_order],
        "硬点大于7g/10g（次数）": [int(hardpoint_counts.get(s, 0)) for s in station_order],
        "磨耗大于8mm（次数）":   [int(wear_counts.get(s, 0)) for s in station_order],
    })


def add_total_row(df: pd.DataFrame, label_col: str) -> pd.DataFrame:
    """在末尾追加合计行"""
    total = {col: int(df[col].sum()) if col != label_col else "合计" for col in df.columns}
    return pd.concat([df, pd.DataFrame([total])], ignore_index=True)


# ---------------------------------------------------------------------------
# Excel 写入
# ---------------------------------------------------------------------------

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_header(ws, row: int, col: int, value: str):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True)
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER


def _style_cell(ws, row: int, col: int, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER


def write_table1_excel(df: pd.DataFrame, outpath: str, direction: str):
    wb = Workbook()
    ws = wb.active
    ws.title = f"拉出值区间统计_{direction}"

    for c, col_name in enumerate(df.columns, 1):
        _style_header(ws, 1, c, col_name)

    for r, (_, row) in enumerate(df.iterrows()):
        for c, val in enumerate(row, 1):
            _style_cell(ws, r + 2, c, int(val) if isinstance(val, (np.integer,)) else val)

    col_widths = {"A": 16, "B": 14, "C": 28, "D": 24, "E": 22}
    for letter, w in col_widths.items():
        ws.column_dimensions[letter].width = w
    ws.freeze_panes = "A2"

    os.makedirs(os.path.dirname(outpath), exist_ok=True)
    wb.save(outpath)
    print(f"已保存: {outpath}")


def write_table2_excel(df: pd.DataFrame, outpath: str, direction: str):
    wb = Workbook()
    ws = wb.active
    ws.title = f"站点统计_{direction}"

    for c, col_name in enumerate(df.columns, 1):
        _style_header(ws, 1, c, col_name)

    for r, (_, row) in enumerate(df.iterrows()):
        for c, val in enumerate(row, 1):
            _style_cell(ws, r + 2, c, int(val) if isinstance(val, (np.integer,)) else val)

    col_widths = {"A": 24, "B": 28, "C": 24, "D": 22}
    for letter, w in col_widths.items():
        ws.column_dimensions[letter].width = w
    ws.freeze_panes = "A2"

    os.makedirs(os.path.dirname(outpath), exist_ok=True)
    wb.save(outpath)
    print(f"已保存: {outpath}")


# ---------------------------------------------------------------------------
# 主函数
# ---------------------------------------------------------------------------

def run(data_dir: str = "original_data",
        output_dir: str = "output_excel",
        station_json_path: str = "station.json",
        line_number: str = ""):
    """供 Web 服务调用：从 data_dir 加载最新日期的上下行文件，

    生成 4 个汇总表格并写入 output_dir：
      - summary_table1_pullout_上行.xlsx / summary_table1_pullout_下行.xlsx
      - summary_table2_station_上行.xlsx / summary_table2_station_下行.xlsx

    所有路径均已参数化，支持多用户并发。
    line_number: 从 welcome.html 用户交互获得的线路号。
    """
    up_path, down_path, date_str = _find_latest_files(data_dir)

    df_up, df_down = pd.DataFrame(), pd.DataFrame()

    if up_path:
        df_up = pd.read_excel(up_path)
        before = len(df_up)
        df_up = df_up.dropna(subset=["站区"])
        if len(df_up) < before:
            print(f"  已过滤 {before - len(df_up)} 行无效数据（站区为空）")
        print(f"  加载 {os.path.basename(up_path)}: {len(df_up)} 行 (上行)")

    if down_path:
        df_down = pd.read_excel(down_path)
        before = len(df_down)
        df_down = df_down.dropna(subset=["站区"])
        if len(df_down) < before:
            print(f"  已过滤 {before - len(df_down)} 行无效数据（站区为空）")
        print(f"  加载 {os.path.basename(down_path)}: {len(df_down)} 行 (下行)")

    for direction, df in [("上行", df_up), ("下行", df_down)]:
        print(f"\n{'=' * 50}")
        print(f"处理 {direction} 数据 ({len(df)} 行)")
        print(f"{'=' * 50}")

        # 数据表1
        table1 = build_table1(df, line_number=line_number)
        table1 = add_total_row(table1, "拉出值区间")
        print(f"\n数据表1（拉出值区间统计 - {direction}）:")
        print(table1.head(5).to_string(index=False))
        print("  ...")
        print(table1.tail(3).to_string(index=False))

        out1 = os.path.join(output_dir, f"summary_table1_pullout_{direction}.xlsx")
        write_table1_excel(table1, out1, direction)

        # 数据表2
        station_order = _build_ordered_station_list(df, direction,
                                                      station_json_path,
                                                      line_number=line_number)
        table2 = build_table2(df, station_order, line_number=line_number)
        table2 = add_total_row(table2, "站点")
        print(f"\n数据表2（站点统计 - {direction}）:")
        print(table2.head(5).to_string(index=False))
        print("  ...")
        print(table2.tail(3).to_string(index=False))

        out2 = os.path.join(output_dir, f"summary_table2_station_{direction}.xlsx")
        write_table2_excel(table2, out2, direction)

    print(f"\n全部表格生成完毕！")


def main():
    run("original_data", "output_excel", "station.json")


if __name__ == "__main__":
    main()
